"""
FRENTE 3 · FECHO — Excel formatado de saída
Módulo 2 — Automação de Documentos

A terceira saída do MESMO esqueleto: devolver a planilha já limpa e
apresentável, sem formatar nada à mão. Repare: ler é igual; muda só a
ferramenta de saída (aqui, openpyxl) (slide 10 e 13).

    python3 scripts/frente3_excel_formatado.py

Lê   : saida/servidores_limpo.xlsx   (gerado pela Frente 0)
Grava: saida/servidores_formatado.xlsx  com:
    • cabeçalho com fundo colorido e texto branco
    • filtro automático na linha de cabeçalho
    • painel congelado (primeira linha fixa ao rolar)
    • cores alternadas (zebra) nas linhas
    • largura das colunas ajustada ao conteúdo
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent.parent
PLANILHA = RAIZ / "saida" / "modulo2" / "servidores_limpo.xlsx"
SAIDA = RAIZ / "saida" / "modulo2" / "servidores_formatado.xlsx"

# Paleta institucional (cores em formato hexadecimal ARGB do openpyxl).
COR_CABECALHO = "FF1F4E79"   # azul escuro
COR_TEXTO_CABECALHO = "FFFFFFFF"  # branco
COR_ZEBRA = "FFD9E1F2"       # azul bem claro para as linhas pares


def estilizar(ws, n_linhas: int, n_colunas: int) -> None:
    """Aplica todo o visual. Cada bloco é um requisito do slide 13 (decomposição)."""
    cabecalho_fill = PatternFill("solid", fgColor=COR_CABECALHO)
    cabecalho_fonte = Font(bold=True, color=COR_TEXTO_CABECALHO)
    zebra_fill = PatternFill("solid", fgColor=COR_ZEBRA)

    # 1) Cabeçalho colorido com texto branco
    for col in range(1, n_colunas + 1):
        celula = ws.cell(row=1, column=col)
        celula.fill = cabecalho_fill
        celula.font = cabecalho_fonte
        celula.alignment = Alignment(horizontal="center", vertical="center")

    # 2) Cores alternadas (zebra) nas linhas de dados
    for linha in range(2, n_linhas + 1):
        if linha % 2 == 0:
            for col in range(1, n_colunas + 1):
                ws.cell(row=linha, column=col).fill = zebra_fill

    # 3) Filtro automático na linha de cabeçalho
    ultima_col = get_column_letter(n_colunas)
    ws.auto_filter.ref = f"A1:{ultima_col}{n_linhas}"

    # 4) Congelar painel: tudo acima de A2 fica fixo ao rolar
    ws.freeze_panes = "A2"

    # 5) Largura das colunas ajustada ao maior conteúdo (com folga)
    for col in range(1, n_colunas + 1):
        letra = get_column_letter(col)
        maior = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, n_linhas + 1)),
            default=10,
        )
        ws.column_dimensions[letra].width = maior + 3


def main() -> None:
    if not PLANILHA.exists():
        raise SystemExit(
            f"Não achei {PLANILHA.relative_to(RAIZ)}.\n"
            "Rode antes a Frente 0:  python3 scripts/frente0_ler_e_limpar.py"
        )

    df = pd.read_excel(PLANILHA, dtype=str)

    # Monta um Excel novo, célula por célula (assim controlamos o estilo).
    wb = Workbook()
    ws = wb.active
    ws.title = "Servidores"

    # Cabeçalho
    for col, nome_coluna in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col, value=nome_coluna)
    # Dados
    for i, (_, linha) in enumerate(df.iterrows(), start=2):
        for col, nome_coluna in enumerate(df.columns, start=1):
            ws.cell(row=i, column=col, value=linha[nome_coluna])

    estilizar(ws, n_linhas=len(df) + 1, n_colunas=len(df.columns))

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA)
    print(f"[ok] Excel formatado salvo em: {SAIDA.relative_to(RAIZ)}")
    print(f"     {len(df)} servidores, {len(df.columns)} colunas, com zebra,"
          " filtro e painel congelado.")

    # ----------------------------------------------------------------------
    # >>> IDEIAS PARA OS ALUNOS irem além (use os 4 pilares) <<<
    #   • Uma aba por secretaria (decomposição + filtrar antes).
    #   • Coluna calculada "Anos de casa" a partir da Data_Admissao.
    #   • Destacar em vermelho os CPFs marcados como INVÁLIDO.
    #   • Congelar também a primeira COLUNA (Nome) além da primeira linha.
    # ----------------------------------------------------------------------


if __name__ == "__main__":
    main()
